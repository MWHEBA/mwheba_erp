from django.test import TestCase, Client
from django.urls import reverse
from django.contrib.auth import get_user_model
from product.models import Product, Category, Unit
from sale.models import Customer, Sale, SaleItem
from purchase.models import Supplier, Purchase, PurchaseItem
from financial.models import Account, Transaction
from io import BytesIO
import csv
import openpyxl
from decimal import Decimal
from datetime import date, timedelta

User = get_user_model()

class ExportTestCase(TestCase):
    """
    اختبارات وظائف تصدير البيانات في النظام
    """
    def setUp(self):
        """
        إعداد البيانات المطلوبة للاختبارات
        """
        # إنشاء مستخدم للاختبارات
        self.user = User.objects.create_user(
            username='exportuser',
            email='export@example.com',
            password='export123'
        )
        
        # إنشاء عميل HTTP وتسجيل الدخول
        self.client = Client()
        self.client.login(username='exportuser', password='export123')
        
        # إنشاء بيانات الاختبار
        self.category = Category.objects.create(
            name='فئة التصدير',
            description='وصف فئة التصدير',
            created_by=self.user
        )
        
        self.unit = Unit.objects.create(
            name='وحدة التصدير',
            abbreviation='و.ت',
            created_by=self.user
        )
        
        # إنشاء منتجات للاختبار
        for i in range(1, 6):
            Product.objects.create(
                name=f'منتج تصدير {i}',
                code=f'EXP{i:03d}',
                category=self.category,
                unit=self.unit,
                purchase_price=Decimal(f'{i}00.00'),
                sale_price=Decimal(f'{i}50.00'),
                created_by=self.user
            )
        
        # الحصول على المنتجات المنشأة
        self.products = Product.objects.all()
        
        # إنشاء عميل
        self.customer = Customer.objects.create(
            name='عميل التصدير',
            phone='01234567890',
            address='عنوان العميل',
            created_by=self.user
        )
        
        # إنشاء مورد
        self.supplier = Supplier.objects.create(
            name='مورد التصدير',
            phone='09876543210',
            address='عنوان المورد',
            created_by=self.user
        )
        
        # إنشاء حساب مالي
        self.account = Account.objects.create(
            name='حساب التصدير',
            account_type='cash',
            balance=Decimal('10000.00'),
            created_by=self.user
        )
        
        # إنشاء عمليات بيع
        today = date.today()
        
        for i in range(1, 4):
            sale = Sale.objects.create(
                customer=self.customer,
                invoice_no=f'INV-EXP-{i}',
                date=today - timedelta(days=i-1),
                payment_method='cash',
                account=self.account,
                notes=f'بيع للتصدير {i}',
                created_by=self.user
            )
            
            # إضافة عناصر البيع
            SaleItem.objects.create(
                sale=sale,
                product=self.products[i-1],
                quantity=i,
                price=self.products[i-1].sale_price,
                created_by=self.user
            )
    
    def test_export_products_csv(self):
        """
        اختبار تصدير المنتجات بتنسيق CSV
        """
        # استدعاء صفحة تصدير المنتجات بتنسيق CSV
        response = self.client.get(
            reverse('product:export_products'),
            {'format': 'csv'}
        )
        
        # التحقق من نجاح الاستجابة
        self.assertEqual(response.status_code, 200)
        
        # التحقق من نوع المحتوى
        self.assertEqual(response['Content-Type'], 'text/csv')
        
        # التحقق من اسم الملف المرفق
        self.assertIn('attachment; filename=', response['Content-Disposition'])
        self.assertIn('products_', response['Content-Disposition'])
        self.assertIn('.csv', response['Content-Disposition'])
        
        # تحليل محتوى الملف CSV
        content = response.content.decode('utf-8')
        csv_reader = csv.reader(content.splitlines())
        
        # الحصول على العناوين والصفوف
        headers = next(csv_reader)
        rows = list(csv_reader)
        
        # التحقق من العناوين
        expected_headers = ['الكود', 'الاسم', 'الفئة', 'سعر الشراء', 'سعر البيع']
        for header in expected_headers:
            self.assertIn(header, headers)
        
        # التحقق من عدد الصفوف (يجب أن يكون 5 منتجات)
        self.assertEqual(len(rows), 5)
        
        # التحقق من محتوى الصف الأول
        first_row = rows[0]
        self.assertEqual(first_row[headers.index('الكود')], 'EXP001')
        self.assertEqual(first_row[headers.index('الاسم')], 'منتج تصدير 1')
    
    def test_export_products_excel(self):
        """
        اختبار تصدير المنتجات بتنسيق Excel
        """
        # استدعاء صفحة تصدير المنتجات بتنسيق Excel
        response = self.client.get(
            reverse('product:export_products'),
            {'format': 'excel'}
        )
        
        # التحقق من نجاح الاستجابة
        self.assertEqual(response.status_code, 200)
        
        # التحقق من نوع المحتوى
        self.assertEqual(response['Content-Type'], 'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
        
        # التحقق من اسم الملف المرفق
        self.assertIn('attachment; filename=', response['Content-Disposition'])
        self.assertIn('products_', response['Content-Disposition'])
        self.assertIn('.xlsx', response['Content-Disposition'])
        
        # تحليل محتوى ملف Excel
        wb = openpyxl.load_workbook(BytesIO(response.content))
        sheet = wb.active
        
        # التحقق من العناوين
        expected_headers = ['الكود', 'الاسم', 'الفئة', 'سعر الشراء', 'سعر البيع']
        for col, header in enumerate(expected_headers, 1):
            self.assertIn(header, sheet.cell(row=1, column=col).value)
        
        # التحقق من عدد الصفوف (العنوان + 5 منتجات)
        self.assertEqual(sheet.max_row, 6)
        
        # التحقق من محتوى الصف الثاني (المنتج الأول)
        self.assertEqual(sheet.cell(row=2, column=1).value, 'EXP001')
        self.assertEqual(sheet.cell(row=2, column=2).value, 'منتج تصدير 1')
    
    def test_export_sales_csv(self):
        """
        اختبار تصدير المبيعات بتنسيق CSV
        """
        # استدعاء صفحة تصدير المبيعات بتنسيق CSV
        response = self.client.get(
            reverse('sale:export_sales'),
            {'format': 'csv'}
        )
        
        # التحقق من نجاح الاستجابة
        self.assertEqual(response.status_code, 200)
        
        # التحقق من نوع المحتوى
        self.assertEqual(response['Content-Type'], 'text/csv')
        
        # تحليل محتوى الملف CSV
        content = response.content.decode('utf-8')
        csv_reader = csv.reader(content.splitlines())
        
        # الحصول على العناوين والصفوف
        headers = next(csv_reader)
        rows = list(csv_reader)
        
        # التحقق من العناوين
        expected_headers = ['رقم الفاتورة', 'التاريخ', 'العميل', 'المجموع', 'طريقة الدفع']
        for header in expected_headers:
            self.assertIn(header, headers)
        
        # التحقق من عدد الصفوف (يجب أن يكون 3 مبيعات)
        self.assertEqual(len(rows), 3)
    
    def test_export_sales_with_items_excel(self):
        """
        اختبار تصدير المبيعات مع التفاصيل بتنسيق Excel
        """
        # استدعاء صفحة تصدير المبيعات بتنسيق Excel مع التفاصيل
        response = self.client.get(
            reverse('sale:export_sales'),
            {'format': 'excel', 'include_items': 'yes'}
        )
        
        # التحقق من نجاح الاستجابة
        self.assertEqual(response.status_code, 200)
        
        # التحقق من نوع المحتوى
        self.assertEqual(response['Content-Type'], 'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
        
        # تحليل محتوى ملف Excel
        wb = openpyxl.load_workbook(BytesIO(response.content))
        
        # التحقق من عدد الأوراق (يجب أن يكون ورقتين: المبيعات وتفاصيل المبيعات)
        self.assertEqual(len(wb.sheetnames), 2)
        self.assertIn('المبيعات', wb.sheetnames)
        self.assertIn('تفاصيل المبيعات', wb.sheetnames)
        
        # التحقق من محتوى ورقة المبيعات
        sales_sheet = wb['المبيعات']
        self.assertGreaterEqual(sales_sheet.max_row, 4)  # العنوان + 3 مبيعات
        
        # التحقق من محتوى ورقة تفاصيل المبيعات
        items_sheet = wb['تفاصيل المبيعات']
        self.assertGreaterEqual(items_sheet.max_row, 4)  # العنوان + 3 عناصر على الأقل
    
    def test_export_customer_report(self):
        """
        اختبار تصدير تقرير العميل
        """
        # استدعاء صفحة تصدير تقرير العميل
        response = self.client.get(
            reverse('sale:export_customer_report'),
            {'customer_id': self.customer.id, 'format': 'csv'}
        )
        
        # التحقق من نجاح الاستجابة
        self.assertEqual(response.status_code, 200)
        
        # التحقق من نوع المحتوى
        self.assertEqual(response['Content-Type'], 'text/csv')
        
        # تحليل محتوى الملف CSV
        content = response.content.decode('utf-8')
        csv_reader = csv.reader(content.splitlines())
        
        # الحصول على العناوين والصفوف
        headers = next(csv_reader)
        rows = list(csv_reader)
        
        # التحقق من عدد الصفوف (يجب أن يكون 3 مبيعات للعميل)
        self.assertEqual(len(rows), 3)
        
        # التحقق من اسم العميل في الملف
        self.assertIn(self.customer.name, content)
    
    def test_export_product_stock_report(self):
        """
        اختبار تصدير تقرير مخزون المنتجات
        """
        # استدعاء صفحة تصدير تقرير مخزون المنتجات
        response = self.client.get(
            reverse('product:export_stock_report'),
            {'format': 'excel'}
        )
        
        # التحقق من نجاح الاستجابة
        self.assertEqual(response.status_code, 200)
        
        # التحقق من نوع المحتوى
        self.assertEqual(response['Content-Type'], 'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
        
        # تحليل محتوى ملف Excel
        wb = openpyxl.load_workbook(BytesIO(response.content))
        sheet = wb.active
        
        # التحقق من العناوين
        expected_headers = ['كود المنتج', 'اسم المنتج', 'الكمية', 'المستودع']
        for col, header in enumerate(expected_headers, 1):
            self.assertIn(header, sheet.cell(row=1, column=col).value)
    
    def test_export_profit_loss_report(self):
        """
        اختبار تصدير تقرير الأرباح والخسائر
        """
        # تواريخ للتقرير
        today = date.today()
        start_date = today - timedelta(days=7)
        end_date = today
        
        # استدعاء صفحة تصدير تقرير الأرباح والخسائر
        response = self.client.get(
            reverse('financial:export_profit_loss_report'),
            {
                'start_date': start_date.strftime('%Y-%m-%d'),
                'end_date': end_date.strftime('%Y-%m-%d'),
                'format': 'csv'
            }
        )
        
        # التحقق من نجاح الاستجابة
        self.assertEqual(response.status_code, 200)
        
        # التحقق من نوع المحتوى
        self.assertEqual(response['Content-Type'], 'text/csv')
        
        # تحليل محتوى الملف CSV
        content = response.content.decode('utf-8')
        
        # التحقق من وجود العناوين الرئيسية
        self.assertIn('الإيرادات', content)
        self.assertIn('المصروفات', content)
        self.assertIn('صافي الربح', content)
    
    def test_export_financial_transactions(self):
        """
        اختبار تصدير المعاملات المالية
        """
        # إنشاء معاملات مالية للاختبار
        for i in range(3):
            Transaction.objects.create(
                account=self.account,
                transaction_type='income' if i % 2 == 0 else 'expense',
                amount=Decimal(f'{(i+1)*100}.00'),
                date=date.today() - timedelta(days=i),
                description=f'معاملة اختبار التصدير {i+1}',
                created_by=self.user
            )
        
        # استدعاء صفحة تصدير المعاملات المالية
        response = self.client.get(
            reverse('financial:export_transactions'),
            {'format': 'excel'}
        )
        
        # التحقق من نجاح الاستجابة
        self.assertEqual(response.status_code, 200)
        
        # التحقق من نوع المحتوى
        self.assertEqual(response['Content-Type'], 'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
        
        # تحليل محتوى ملف Excel
        wb = openpyxl.load_workbook(BytesIO(response.content))
        sheet = wb.active
        
        # التحقق من العناوين
        expected_headers = ['التاريخ', 'النوع', 'المبلغ', 'الوصف', 'الحساب']
        for col, header in enumerate(expected_headers, 1):
            self.assertIn(header, sheet.cell(row=1, column=col).value)
        
        # التحقق من عدد الصفوف (العنوان + 3 معاملات)
        self.assertEqual(sheet.max_row, 4)
    
    def test_export_with_filtering(self):
        """
        اختبار تصدير البيانات مع الفلترة
        """
        # استدعاء صفحة تصدير المنتجات مع فلترة حسب الفئة
        response = self.client.get(
            reverse('product:export_products'),
            {'format': 'csv', 'category': self.category.id}
        )
        
        # التحقق من نجاح الاستجابة
        self.assertEqual(response.status_code, 200)
        
        # تحليل محتوى الملف CSV
        content = response.content.decode('utf-8')
        csv_reader = csv.reader(content.splitlines())
        
        # الحصول على العناوين والصفوف
        headers = next(csv_reader)
        rows = list(csv_reader)
        
        # التحقق من عدد الصفوف (يجب أن يكون 5 منتجات)
        self.assertEqual(len(rows), 5)
        
        # استدعاء صفحة تصدير المنتجات مع فلترة حسب اسم المنتج
        response = self.client.get(
            reverse('product:export_products'),
            {'format': 'csv', 'name': 'منتج تصدير 1'}
        )
        
        # التحقق من نجاح الاستجابة
        self.assertEqual(response.status_code, 200)
        
        # تحليل محتوى الملف CSV
        content = response.content.decode('utf-8')
        csv_reader = csv.reader(content.splitlines())
        
        # الحصول على العناوين والصفوف
        headers = next(csv_reader)
        rows = list(csv_reader)
        
        # التحقق من عدد الصفوف (يجب أن يكون منتج واحد)
        self.assertEqual(len(rows), 1)
        self.assertEqual(rows[0][headers.index('الاسم')], 'منتج تصدير 1') 